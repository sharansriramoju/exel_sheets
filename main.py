import openpyxl as xl


wb = xl.load_workbook("discount.xlsx")
kc = xl.load_workbook("products.xlsx")
disc_sheet = wb["Sheet1"]
prod_sheet = kc["Sheet1"]
prod_sheet.cell(1, 5).value = 'Discount'
prod_sheet.cell(1, 6).value = 'Bill'
for i in range(2, prod_sheet.max_row+1):
    prod_sheet.cell(i, 5).value = disc_sheet.cell(i, 2).value
    l = prod_sheet.cell(i, 5).value
    prod_sheet.cell(i, 6).value = prod_sheet.cell(i, 3).value*prod_sheet.cell(i, 4).value*(l/100)
kc.save("total.xlsx")
